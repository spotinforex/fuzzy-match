#!/usr/bin/env python3
"""
fuzzy_match.py — Match full names across two Excel sheets using fuzzy matching.

Usage:
    python fuzzy_match.py sheet1.xlsx sheet2.xlsx [OPTIONS]

Options:
    --col1 NAME        Column name in sheet1 (default: auto-detect)
    --col2 NAME        Column name in sheet2 (default: auto-detect)
    --sheet1 NAME      Sheet tab name in file1 (default: first sheet)
    --sheet2 NAME      Sheet tab name in file2 (default: first sheet)
    --threshold INT    Minimum match score 0-100 (default: 70)
    --output FILE      Output filename (default: match_results.xlsx)
    --top INT          Top N matches per name (default: 1)
"""

import argparse
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from thefuzz import fuzz, process


NAME_HINTS = {"name", "full name", "fullname", "names"}


def detect_name_column(df: pd.DataFrame, label: str) -> str:
    for col in df.columns:
        if col.strip().lower() in NAME_HINTS:
            return col
    # Fall back: first string column
    for col in df.columns:
        if df[col].dtype == object:
            return col
    sys.exit(f"[ERROR] Could not detect a name column in {label}. Use --col1/--col2 to specify.")


def normalize(name: str) -> str:
    return " ".join(str(name).strip().split()).upper()


def run_match(names1, names2, threshold, top_n):
    rows = []
    norm2 = [normalize(n) for n in names2]
    choices = {normalize(n): n for n in names2}

    for raw1 in names1:
        n1 = normalize(raw1)
        results = process.extract(n1, list(choices.keys()), scorer=fuzz.token_sort_ratio, limit=top_n)
        for match_key, score in results:
            if score >= threshold:
                rows.append({
                    "Name (Sheet 1)": raw1,
                    "Matched Name (Sheet 2)": choices[match_key],
                    "Match Score (%)": score,
                    "Match Quality": quality_label(score),
                })
            else:
                rows.append({
                    "Name (Sheet 1)": raw1,
                    "Matched Name (Sheet 2)": choices[match_key],
                    "Match Score (%)": score,
                    "Match Quality": "Below Threshold",
                })
    return rows


def quality_label(score: int) -> str:
    if score == 100:
        return "Exact"
    if score >= 90:
        return "Very High"
    if score >= 80:
        return "High"
    if score >= 70:
        return "Medium"
    return "Low"


SCORE_COLORS = {
    "Exact":          "C6EFCE",
    "Very High":      "A9D08E",
    "High":           "FFEB9C",
    "Medium":         "FCE4D6",
    "Low":            "F4CCCC",
    "Below Threshold":"D9D9D9",
}

HDR_FILL  = PatternFill("solid", start_color="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=11, name="Arial")
BODY_FONT = Font(size=10, name="Arial")
CENTER    = Alignment(horizontal="center", vertical="center")
LEFT      = Alignment(horizontal="left",   vertical="center")
THIN      = Side(style="thin", color="BFBFBF")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_output(rows, output_path, threshold):
    wb = Workbook()
    ws = wb.active
    ws.title = "Match Results"

    headers = ["Name (Sheet 1)", "Matched Name (Sheet 2)", "Match Score (%)", "Match Quality"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font    = HDR_FONT
        cell.fill    = HDR_FILL
        cell.alignment = CENTER
        cell.border  = BORDER

    ws.row_dimensions[1].height = 22

    for row_data in rows:
        ws.append([row_data[h] for h in headers])
        row_idx = ws.max_row
        quality = row_data["Match Quality"]
        fill_color = SCORE_COLORS.get(quality, "FFFFFF")
        fill = PatternFill("solid", start_color=fill_color)
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            cell.font   = BODY_FONT
            cell.fill   = fill
            cell.border = BORDER
            cell.alignment = CENTER if col_idx in (3, 4) else LEFT
        ws.row_dimensions[row_idx].height = 18

    col_widths = [35, 35, 18, 16]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{ws.max_row}"

    # ── Summary sheet ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    total      = len(rows)
    above      = sum(1 for r in rows if r["Match Quality"] != "Below Threshold")
    exact      = sum(1 for r in rows if r["Match Quality"] == "Exact")
    below      = total - above

    summary_rows = [
        ["Metric", "Count"],
        ["Total names processed", total],
        ["Matches above threshold", above],
        ["Exact matches (100%)", exact],
        ["Below threshold", below],
        ["Threshold used", f"{threshold}%"],
    ]
    for sr in summary_rows:
        ws2.append(sr)

    for cell in ws2[1]:
        cell.font   = HDR_FONT
        cell.fill   = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        for cell in row:
            cell.font      = BODY_FONT
            cell.alignment = LEFT
            cell.border    = BORDER

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 18

    wb.save(output_path)
    return above, total


def main():
    parser = argparse.ArgumentParser(
        description="Fuzzy-match full names between two Excel sheets.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("file1",       help="First Excel file")
    parser.add_argument("file2",       help="Second Excel file")
    parser.add_argument("--col1",      help="Column name in file1")
    parser.add_argument("--col2",      help="Column name in file2")
    parser.add_argument("--sheet1",    help="Sheet tab in file1 (default: first)")
    parser.add_argument("--sheet2",    help="Sheet tab in file2 (default: first)")
    parser.add_argument("--threshold", type=int, default=70, metavar="0-100",
                        help="Minimum match score (default: 70)")
    parser.add_argument("--output",    default="match_results.xlsx",
                        help="Output file (default: match_results.xlsx)")
    parser.add_argument("--top",       type=int, default=1, metavar="N",
                        help="Top N matches per name (default: 1)")
    args = parser.parse_args()

    if not (0 <= args.threshold <= 100):
        sys.exit("[ERROR] --threshold must be between 0 and 100.")

    print(f"Loading files...")
    try:
        df1 = pd.read_excel(args.file1, sheet_name=args.sheet1 or 0)
        df2 = pd.read_excel(args.file2, sheet_name=args.sheet2 or 0)
    except FileNotFoundError as e:
        sys.exit(f"[ERROR] {e}")

    col1 = args.col1 or detect_name_column(df1, "file1")
    col2 = args.col2 or detect_name_column(df2, "file2")

    if col1 not in df1.columns:
        sys.exit(f"[ERROR] Column '{col1}' not found in file1. Available: {list(df1.columns)}")
    if col2 not in df2.columns:
        sys.exit(f"[ERROR] Column '{col2}' not found in file2. Available: {list(df2.columns)}")

    names1 = df1[col1].dropna().astype(str).tolist()
    names2 = df2[col2].dropna().astype(str).tolist()

    print(f"Sheet 1: {len(names1)} names  (column: '{col1}')")
    print(f"Sheet 2: {len(names2)} names  (column: '{col2}')")
    print(f"Threshold: {args.threshold}%  |  Top matches per name: {args.top}")
    print(f"Running fuzzy match...")

    rows = run_match(names1, names2, args.threshold, args.top)
    above, total = write_output(rows, args.output, args.threshold)

    print(f"Done! Results saved to: {args.output}")
    print(f"{above}/{total} matches at or above {args.threshold}% threshold\n")


if __name__ == "__main__":
    main()